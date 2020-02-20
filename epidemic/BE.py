import pymysql
import datetime
import openpyxl
import os
import sys


class Database:
    def __init__(self):
        self.conn = pymysql.connect(
            host="localhost", user="root", password="zhouenhua", database="epidemic", charset="utf8")
        self.results = dict()

    def get_all(self):
        with self.conn:
            cursor = self.conn.cursor()
            sql = 'select * from patient'
            cursor.execute(sql)

            results = cursor.fetchall()

            return results

    def get_today_provenience(self, date):
        with self.conn:

            results = []
            cursor = self.conn.cursor()

            sql_provenience_num = "select provenience,count(*) num from patient where observeStartDate = %s group by provenience"
            cursor.execute(sql_provenience_num, date)

            for result in cursor.fetchall():
                results.append(result)
            self.results.update(results)

    def get_today_statistic(self, date):
        with self.conn:

            cursor = self.conn.cursor(pymysql.cursors.DictCursor)

            sql_today_total_observe = 'select count(*) today_total_observe from patient where observeStartDate <= %s'
            cursor.execute(sql_today_total_observe, date)
            self.results.update(cursor.fetchone())

            sql_today_add = "select count(*) today_add from patient where observeStartDate = %s"
            cursor.execute(sql_today_add, date)
            self.results.update(cursor.fetchone())

            sql_today_total_relieve = "select count(*) today_total_relieve from patient where observeEndDate <= %s"
            cursor.execute(sql_today_total_relieve, date)
            self.results.update(cursor.fetchone())

            sql_today_relieve = "select count(*) today_relieve from patient where observeEndDate = %s"
            cursor.execute(sql_today_relieve, date)
            self.results.update(cursor.fetchone())

            sql_today_still_observe = "select count(*) today_still_observe from patient where observeEndDate > %s and observeStartDate <= %s"
            cursor.execute(sql_today_still_observe, [date, date])
            self.results.update(cursor.fetchone())

    def get_yesterday_statistic(self, date):
        with self.conn:
            cursor = self.conn.cursor(pymysql.cursors.DictCursor)

            sql_yesterday_total_observe = "select count(*) yesterday_total_observe from patient where observeStartDate <= %s"
            cursor.execute(sql_yesterday_total_observe, date)
            self.results.update(cursor.fetchone())

            sql_yesterday_total_relieve = "select count(*) yesterday_total_relieve from patient where observeEndDate <= %s"
            cursor.execute(sql_yesterday_total_relieve, date)
            self.results.update(cursor.fetchone())

            sql_yesterday_still_observe = "select count(*) yesterday_still_observe from patient where observeEndDate > %s and observeStartDate <= %s"
            cursor.execute(sql_yesterday_still_observe, [date, date])
            self.results.update(cursor.fetchone())

    def insert_into(self, data):
        with self.conn:
            cursor = self.conn.cursor()
            sql = "replace into patient (idNo,name,sex,age,address,phone,provenience,observeStartDate,ObserveEndDate,isContactBefore125,isContactAfter125,isSymptom,isPSBfocus,functionaryMaster,functionary,note) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

            cursor.executemany(sql, data)
            self.conn.commit()


class ReadExcel:

    def __init__(self, file_name, sheet_name):
        self.wb = openpyxl.load_workbook(file_name)
        self.sh = self.wb[sheet_name]

    def read_excel(self):

        data = []
        note_cach = ""

        for row in self.sh.iter_rows(min_row=6):
            patient = []
            idNo = row[4].value

            age = self.get_age(idNo)
            sex = self.get_sex(idNo)
            startDate = self.get_start_date(row[8].value)
            endDate = startDate + datetime.timedelta(days=14)

            # idNo
            patient.append(idNo)
            # name
            patient.append(row[1].value)
            # sex
            patient.append(sex)
            # age
            patient.append(age)
            # address
            patient.append(row[5].value)
            # phone
            patient.append(row[6].value)
            # provenience
            patient.append(row[7].value)
            # observeStartDate
            patient.append(startDate)
            # observeEndDate
            patient.append(endDate)
            # isContactBefore125
            patient.append(self.char_to_boolean(row[11].value))
            # isContactAfter125
            patient.append(self.char_to_boolean(row[15].value))
            # isSymptom
            patient.append(self.char_to_boolean(row[18].value))
            # isPSBfocus
            patient.append(self.char_to_boolean(row[19].value))
            # functionaryMaster
            patient.append(row[20].value)
            # functionary
            patient.append(row[21].value)
            # note
            if(row[22].value != None):
                note_cach = row[22].value
                patient.append(row[22].value)
            else:
                patient.append(note_cach)
            data.append(patient)
        return data

    def get_age(self, idNo):
        now = (datetime.datetime.now() + datetime.timedelta(days=1))
        year = now.year
        month = now.month
        day = now.day
        birth_year = int(idNo[6:10])
        birth_month = int(idNo[10:12])
        birth_day = int(idNo[12:14])

        if year == birth_year:
            return 0
        else:
            if birth_month > month or(birth_month == month and birth_day > day):
                return year - birth_year - 1
            else:
                return year - birth_year

    def get_sex(self, idNo):
        num = int(idNo[16:17])
        if num % 2 == 0:
            return "女"
        else:
            return "男"

    def get_start_date(self, startDate):
        month = int(startDate[0:2])
        day = int(startDate[3:5])
        return datetime.date(2020, month, day)

    def char_to_boolean(self, char):
        if char == "是" or char == "有":
            return 1
        else:
            return 0


class ExportExcel:

    def __init__(self, sheet_name):
        self.wb = openpyxl.load_workbook(os.path.dirname(os.path.realpath(sys.argv[0])) + '\\module\\模板文件.xlsx')
        self.sh = self.wb[sheet_name]

    def export_excel(self, data, seldate, filepath):

        StrTodayDate = '%s月%s日' % (str(seldate.month), str(seldate.day))
        StrYesterdayDate = '%s月%s日' % (
            str(seldate.month), str(seldate.day - 1))

        self.sh['R2'] = '数据统计时间：%s12:00至%s12:00' % (
            StrYesterdayDate, StrTodayDate)
        self.sh['C3'] = StrTodayDate + '情况'
        self.sh['U3'] = StrYesterdayDate + '情况'

        export_data = self.handle_provenience_data(data)

        for i in range(23):
            self.sh.cell(row=8, column=i + 1, value=export_data[i])

        self.save_filename = filepath + '\\' + str(seldate) + '.xlsx'

        self.wb.save(self.check_filename_available(self.save_filename))

    def handle_provenience_data(self, data):
        return_data = ['' for i in range(23)]

        return_data[0] = '1'
        return_data[1] = '双土'
        return_data[2] = data.get('today_total_observe')
        return_data[3] = data.get('today_add')

        return_data[4] = data.get('湖北武汉', 0)

        return_data[6] = data.get('广东', 0)
        return_data[7] = data.get('浙江', 0)
        return_data[8] = data.get('河南', 0)
        return_data[9] = data.get('湖南', 0)
        return_data[10] = data.get('安徽', 0)
        return_data[11] = data.get('江西', 0)
        return_data[12] = data.get('江苏', 0)

        return_data[17] = data.get('today_total_relieve')
        return_data[18] = data.get('today_relieve')
        return_data[19] = data.get('today_still_observe')

        return_data[20] = data.get('yesterday_total_observe')
        return_data[21] = data.get('yesterday_total_relieve')
        return_data[22] = data.get('yesterday_still_observe')

        hubei_all = 0
        chongqing_all = 0
        wanzou_all = 0
        yunyang_all = 0

        for key in data:
            if '湖北' in key:
                hubei_all += data.get(key,0)
            if '重庆' in key:
                chongqing_all += data.get(key,0)
            if '万州' in key:
                wanzou_all += data.get(key,0)
            if '云阳' in key:
                yunyang_all += data.get(key,0)

        return_data[5] = hubei_all - return_data[4]
        return_data[14] = wanzou_all
        return_data[15] = chongqing_all - wanzou_all - yunyang_all
        return_data[16] = yunyang_all

        other_num = return_data[3]

        for i in return_data[4:13]:
            other_num -= i
        
        for i in return_data[14:17]:
            other_num -= i

        return_data[13] = other_num 

        return return_data

    def check_filename_available(self, filename):
        n = [1]

        def check_meta(file_name):
            file_name_new = file_name
            if os.path.isfile(file_name):
                file_name_new = file_name[:file_name.rfind(
                    '.')] + ' (' + str(n[0]) + ')' + file_name[file_name.rfind('.'):]
                n[0] += 1
            if os.path.isfile(file_name_new):
                file_name_new = check_meta(file_name)
            return file_name_new

        return_name = check_meta(filename)

        return return_name


class Business:
    def __init__(self):
        self.database = Database()

    def import_data(self, filename):
        readExcel = ReadExcel(filename, "Sheet1")
        sqldata = readExcel.read_excel()

        self.database.insert_into(sqldata)

    def search_data(self, date):

        self.database.get_today_statistic(date)
        self.database.get_today_provenience(date)
        self.database.get_yesterday_statistic(
            date - datetime.timedelta(days=1))

        return self.database.results

    def export_data(self, data, filepath, seldate):

        exportExcel = ExportExcel('Sheet1')

        exportExcel.export_excel(data, seldate, filepath)


# database = Database()

# database.insert_into(sqldata)
# result = database.get_all()
# database.get_statistic("2020-02-14")
# result = database.get_today_statistic('2020-02-01')
# result = database.get_yesterday_statistic('2020-02-05')
# result = database.get_today_provenience('2020-02-06')
